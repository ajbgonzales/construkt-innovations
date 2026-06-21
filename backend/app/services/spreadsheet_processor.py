import openpyxl
import pandas as pd

from datetime import datetime, timedelta

from io import BytesIO

from models.attendance import EmployeeAttendanceRecord

from services.dates import get_date_range
from services.time_logs import get_time_logs


def get_metadata(projects_metadata: dict):
    project_name = projects_metadata["project_name"]
    start_time = projects_metadata["start_time"]
    is_compressed = projects_metadata["is_compressed"]
    is_overtime = projects_metadata["is_overtime"]
    # working_days = projects_metadata["working_days"]

    return project_name, start_time, is_compressed, is_overtime


def clean_attendance_spreadsheet(df: pd.DataFrame, projects_metadata: dict):
    project_name, start_time, is_compressed, is_overtime = get_metadata(
        projects_metadata
    )

    df.columns = [
        f"col_{i}" if col.startswith("Unnamed") else col
        for i, col in enumerate(df.columns, start=1)
    ]

    start_date, end_date = get_date_range(df.loc[1, "col_11"])
    records = get_employee_records(
        df, project_name, start_date, end_date, is_compressed
    )
    create_cleaned_spreadsheet(records, project_name)


def get_employee_records(
    df: pd.DataFrame,
    project: str,
    start_date: datetime,
    end_date: datetime,
    is_compressed_time: bool,
):
    records: list[EmployeeAttendanceRecord] = []
    rows = list(df.itertuples(index=False))
    for i, row in enumerate(rows[:-1]):
        if row.col_4 == "User ID:":
            current = start_date
            col_num = 1
            while current <= end_date:
                time_in, time_out, is_flagged, notes = get_time_logs(
                    rows,
                    i + 2,
                    current,
                    col_num,
                    is_compressed_time,
                )
                record = EmployeeAttendanceRecord(
                    employee_id=row.col_5,
                    employee_full_name=row.col_8,
                    position=row.col_10,
                    project=project,
                    rate=0,
                    allowance=0,
                    phic=0,
                    hdmf=0,
                    date=current,
                    time_in=time_in,
                    time_out=time_out,
                    break_seconds=1800 if is_compressed_time else 3600,
                    is_compressed_time=is_compressed_time,
                    is_flagged=is_flagged,
                    notes=notes,
                )
                records.append(record)
                current += timedelta(days=1)
                col_num += 1

    return records


def create_cleaned_spreadsheet(
    records: list[EmployeeAttendanceRecord], project_name: str
):
    cleaned_dict = {}
    for r in records:
        # Create a new row for the employee if not yet present
        if r.employee_id not in cleaned_dict:
            cleaned_dict[r.employee_id] = {
                "Employee ID": r.employee_id,
                "Employee Full Name": r.employee_full_name,
                "Position": r.position,
                "Project": r.project,
                "Rate": r.rate,
                "Allowance": r.allowance,
                "PHIC": r.phic,
                "HDMF": r.hdmf,
                "Is Flagged": "No",
                "Notes": None,
                "Total Work Hours": 0,
            }
        # Flag employee
        if cleaned_dict[r.employee_id]["Is Flagged"] == "No" and r.is_flagged == "Yes":
            cleaned_dict[r.employee_id]["Is Flagged"] = r.is_flagged
        # Add notes
        if r.notes:
            if cleaned_dict[r.employee_id]["Notes"]:
                cleaned_dict[r.employee_id]["Notes"] += f"\n{r.notes}"
            else:
                cleaned_dict[r.employee_id]["Notes"] = r.notes
        # Add work hours
        if r.time_in and r.time_out:
            hours = ((r.time_out - r.time_in).total_seconds() - r.break_seconds) / 3600
            cleaned_dict[r.employee_id][r.date.strftime("%Y-%m-%d")] = hours
            cleaned_dict[r.employee_id]["Total Work Hours"] += hours
        else:
            cleaned_dict[r.employee_id][r.date.strftime("%Y-%m-%d")] = 0

    # Convert cleaned_dict to data frame
    cleaned_df = pd.DataFrame.from_dict(cleaned_dict, orient="index")

    # Move Total Work Hours column to last position
    total_work_hours_col = cleaned_df.pop("Total Work Hours")
    cleaned_df["Total Work Hours"] = total_work_hours_col

    # Convert data frame to excel
    cleaned_df.to_excel(
        f"./app/records/{project_name}.xlsx", sheet_name=f"{project_name}", index=False
    )


def compile_spreadsheets(file_paths: list[str], buffer: BytesIO):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for path in file_paths:
        wb = openpyxl.load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_ws = workbook.create_sheet(title=sheet_name)
            for row in ws.iter_rows(values_only=True):
                new_ws.append(row)

    workbook.save(buffer)
    buffer.seek(0)
    return buffer
